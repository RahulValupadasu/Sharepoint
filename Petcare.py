# Databricks notebook source
import os
from io import BytesIO
from typing import Optional

import pandas as pd
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.user_credential import UserCredential
from office365.sharepoint.files.file import File
from office365.runtime.client_request_exception import ClientRequestException
from pyspark.sql import SparkSession
from urllib.parse import quote


# COMMAND ----------

# SharePoint configuration
sp_url = "https://zoetis.sharepoint.com/"
# full server-relative path to the SharePoint site
site_path = "sites/PetcareBrandMarketingTeam"
site_url = f"{sp_url}{site_path}"
# document library name (spaces do not need encoding)
document_library = "Shared Documents"

# File relative path inside the document library
orders_summary_relative_path = (
     "Core Brands/Promotions/Vanguard Annual Programs/2025/"
    "2025 CAC - free doses/Tracker/Orders summary.xlsx"
)

# Placeholder path in ADLS where the Excel file will be persisted
adls_output_path = (
    "abfss://<container>@<storage-account>.dfs.core.windows.net/path/"
    "OrdersSummary.xlsx"
)

# Credentials from Databricks secrets
# "dbutils.secrets" retrieves secrets stored in Databricks. Replace these
# lines with your own secrets management solution when running elsewhere.
scope = os.environ.get("NGSE_KEY_VAULT_SCOPE")
username = dbutils.secrets.get(scope=scope, key="svc-azr-ngsesharepnt-user")
password = dbutils.secrets.get(scope=scope, key="svc-azr-ngsesharepnt-password")

credentials = UserCredential(username, password)
ctx = ClientContext(site_url).with_credentials(credentials)

# COMMAND ----------

def check_connection(context: ClientContext) -> None:
    """Validate SharePoint connection."""
    try:
        context.web.get().execute_query()
        print(f"Connected to site: {site_url}")
    except ClientRequestException as exc:
        if exc.response_code in (401, 403):
            raise PermissionError("Unauthorized access to SharePoint") from exc
        raise


# COMMAND ----------

import os
import uuid
from office365.sharepoint.files.file import File
from office365.runtime.client_request_exception import ClientRequestException
from pyspark.sql import DataFrame
import pyspark.sql
from openpyxl import load_workbook

def read_excel_as_spark(
    context: ClientContext,
    library: str,
    relative_path: str,
    sheet_name: str = "Sheet1",
) -> pyspark.sql.DataFrame:
    """
    Download an Excel file from SharePoint then read it via spark-excel.
    """
    # 1) server-relative URL
    file_url = f"/{site_path}/{library}/{relative_path}"
    print(f"Downloading: {file_url}")

    # 2) fetch bytes
    try:
        resp = File.open_binary(context, file_url)
        content = resp.content
    except ClientRequestException as exc:
        if exc.response_code in (401, 403):
            raise PermissionError("Unauthorized") from exc
        if exc.response_code == 404:
            raise FileNotFoundError(f"Not found: {file_url}") from exc
        raise

    # 3) write to a temporary local file
    tmp_filename = f"/tmp/{uuid.uuid4()}.xlsx"
    with open(tmp_filename, "wb") as f:
        f.write(content)
    # Convert the file to non-strict OOXML format
    wb = load_workbook(tmp_filename)
    
    # Ensure at least one sheet is visible
    if not any(sheet.sheet_state == "visible" for sheet in wb.worksheets):
        if wb.worksheets:
            wb.worksheets[0].sheet_state = "visible"
    
    wb.save(tmp_filename)
    spark_path = f"file://{tmp_filename}"

    # 4) read as Spark DataFrame
    df = (
        spark.read
             .format("com.crealytics.spark.excel")
             .option("sheetName", sheet_name)      # which sheet to load
             .option("header", "true")             # first row as header
             .option("inferSchema", "true")        # auto-detect dtypes
             .option("treatEmptyValuesAsNulls", "true")
             .load(spark_path)
    )
    return df


def save_excel_to_adls(
    context: ClientContext,
    library: str,
    relative_path: str,
    adls_path: str,
) -> None:
    """Download an Excel file from SharePoint and copy it to ADLS."""
    file_url = f"/{site_path}/{library}/{relative_path}"
    print(f"Downloading for ADLS: {file_url}")

    try:
        resp = File.open_binary(context, file_url)
        content = resp.content
    except ClientRequestException as exc:
        if exc.response_code in (401, 403):
            raise PermissionError("Unauthorized") from exc
        if exc.response_code == 404:
            raise FileNotFoundError(f"Not found: {file_url}") from exc
        raise

    tmp_filename = f"/tmp/{uuid.uuid4()}.xlsx"
    try:
        with open(tmp_filename, "wb") as f:
            f.write(content)

        wb = load_workbook(tmp_filename)
        if not any(sheet.sheet_state == "visible" for sheet in wb.worksheets):
            if wb.worksheets:
                wb.worksheets[0].sheet_state = "visible"
        wb.save(tmp_filename)

        dbutils.fs.cp(f"file:{tmp_filename}", adls_path, True)
    finally:
        dbutils.fs.rm(f"file:{tmp_filename}")

# Usage:
check_connection(ctx)
save_excel_to_adls(ctx, document_library, orders_summary_relative_path, adls_output_path)
spark_df = read_excel_as_spark(ctx, document_library, orders_summary_relative_path, sheet_name="Sheet1")
display(spark_df)
